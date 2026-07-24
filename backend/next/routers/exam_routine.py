from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import pandas as pd
import io
import logging
from database import get_raw_db

logger = logging.getLogger("uvicorn.error")

router = APIRouter(
    prefix="/api/routines",
    tags=["Exam Routines"]
)


# ----------------------------
# GET ALL ROUTINES (teacher/student dashboards)
# ----------------------------
@router.get("/all")
async def get_all_routines():
    try:
        with get_raw_db() as conn:
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT
                    batch_name,
                    exam_date,
                    exam_time,
                    subject_name,
                    subject_code
                FROM exam_routines
                ORDER BY batch_name ASC, exam_date ASC, exam_time ASC;
                """
            )

            records = cursor.fetchall()
            results = []

            for row in records:
                results.append({
                    "batch_name": row["batch_name"],
                    "exam_date": str(row["exam_date"]),
                    "exam_time": row["exam_time"] or "",
                    "subject_name": row["subject_name"],
                    "subject_code": row["subject_code"],
                })

            return results

    except Exception as e:
        logger.error(str(e))
        raise HTTPException(status_code=500, detail=str(e))


# ----------------------------
# GET ROUTINE BY BATCH
# ----------------------------
@router.get("")
async def get_routines(batch: str):
    try:
        with get_raw_db() as conn:
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT
                    exam_date,
                    exam_time,
                    subject_name,
                    subject_code
                FROM exam_routines
                WHERE batch_name=%s
                ORDER BY exam_date ASC;
                """,
                (batch,)
            )

            records = cursor.fetchall()

            results = []

            for row in records:
                results.append({
                    "Date": str(row["exam_date"]),
                    "Time": row["exam_time"] if row["exam_time"] else "",
                    "Subject": row["subject_name"],
                    "Code": row["subject_code"]
                })

            cursor.close()
            return results

    except Exception as e:
        logger.error(str(e))
        raise HTTPException(status_code=500, detail=str(e))


# ----------------------------
# UPLOAD ROUTINE
# ----------------------------
def normalize_time(time_val):
    """
    Normalize exam_time: strip whitespace and standardize spaces around dashes
    so "10:00 AM - 12:00 PM" and "10:00 AM-12:00 PM" become the same slot.
    Also handles Excel serial dates (pd.Timestamp) that appear when Excel
    stores a time-only cell as a datetime.
    """
    if pd.isna(time_val):
        return ""
    
    # Handle Excel serial date / pd.Timestamp (e.g. 1899-12-30 10:00:00)
    if isinstance(time_val, pd.Timestamp):
        # Extract just the time portion as HH:MM AM/PM
        return time_val.strftime("%I:%M %p").lstrip("0")
    
    s = str(time_val).strip()
    if not s:
        return ""
    
    import re
    
    # Detect if this is a date-like string (e.g. "1899-12-30" with or without a time)
    date_match = re.match(r'^\d{4}-\d{2}-\d{2}(\s|T)(\d{2}:\d{2})', s)
    if date_match:
        # Extract the time part: HH:MM -> convert to 12-hour with AM/PM
        time_part = date_match.group(2)
        try:
            hours, minutes = time_part.split(":")
            h = int(hours)
            ampm = "AM" if h < 12 else "PM"
            if h == 0:
                h = 12
            elif h > 12:
                h -= 12
            return f"{h}:{minutes} {ampm}"
        except:
            pass
    
    # Normalize spaces around hyphens/dashes: "10:00 - 12:00" -> "10:00-12:00"
    s = re.sub(r'\s*-\s*', '-', s)
    
    # Normalize spaces after AM/PM: "10:00AM" -> "10:00 AM"
    s = re.sub(r'(\d)([AaPp][Mm])', r'\1 \2', s)
    
    # Uppercase AM/PM so "am" / "pm" both become "AM" / "PM"
    s = re.sub(r'([AaPp][Mm])', lambda m: m.group(1).upper(), s)
    
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s)
    
    return s


@router.post("/bulk")
async def upload_routine(
    file: UploadFile = File(...),
    batch: str = Form(...)
):

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Upload an Excel file."
        )

    try:

        contents = await file.read()

        df = pd.read_excel(io.BytesIO(contents))

        df.columns = [c.strip().capitalize() for c in df.columns]

        required = ["Date", "Subject", "Code"]

        for col in required:
            if col not in df.columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing column '{col}'"
                )

        # Optional time column
        if "Time" not in df.columns:
            df["Time"] = ""

        # ----------------------------
        # Convert Excel dates
        # ----------------------------
        def convert_excel_date(value):

            if pd.isna(value):
                return None

            # Excel serial number
            if isinstance(value, (int, float)):
                return pd.to_datetime(
                    value,
                    unit="D",
                    origin="1899-12-30"
                ).date()

            # Already a datetime
            if isinstance(value, pd.Timestamp):
                return value.date()

            # Normal string
            return pd.to_datetime(value).date()

        df["Date"] = df["Date"].apply(convert_excel_date)
        
        # Normalize the time column
        df["Time"] = df["Time"].apply(normalize_time)

        with get_raw_db() as conn:

            cursor = conn.cursor()

            try:

                cursor.execute(
                    """
                    DELETE FROM exam_routines
                    WHERE batch_name=%s;
                    """,
                    (batch,)
                )

                for _, row in df.iterrows():
                    cursor.execute(
                        """
                        INSERT INTO exam_routines
                        (
                            batch_name,
                            exam_date,
                            subject_name,
                            subject_code,
                            exam_time
                        )

                        VALUES
                        (%s,%s,%s,%s,%s);
                        """,
                        (
                            batch,
                            str(row["Date"]),
                            row["Subject"],
                            row["Code"],
                            str(row["Time"]) if row["Time"] else ""
                        )
                    )

                conn.commit()

            except Exception as e:
                conn.rollback()
                raise e

            finally:
                cursor.close()

        return {
            "message": "Routine uploaded successfully."
        }

    except Exception as e:
        logger.error(str(e))
        raise HTTPException(status_code=400, detail=str(e))
@router.delete("/batch/{batch_name}")
async def delete_routines_by_batch(batch_name: str):
    try:
        with get_raw_db() as conn:
            cursor = conn.cursor()

            cursor.execute(
                """
                DELETE FROM exam_routines
                WHERE batch_name=%s;
                """,
                (batch_name,)
            )

            deleted_count = cursor.rowcount
            conn.commit()
            cursor.close()

        return {
            "message": f"Deleted {deleted_count} routine entries for batch '{batch_name}'."
        }

    except Exception as e:
        logger.error(str(e))
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/template")
async def download_template():
    """Download a sample Excel template with the exact columns required."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Routine Template"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="5668F5", end_color="5668F5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Date", "Subject", "Code", "Time"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border

    sample_data = ["2026-08-15", "Mathematics", "MATH101", "10:00 AM - 12:00 PM"]
    for col_idx, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 25; ws.column_dimensions["C"].width = 15; ws.column_dimensions["D"].width = 30

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exam_routine_template.xlsx"})


@router.get("/sessions")
def get_sessions():
    with get_raw_db() as conn:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT
                exam_date,
                exam_time
            FROM exam_routines
            ORDER BY exam_date, exam_time;
        """)

        return cursor.fetchall()